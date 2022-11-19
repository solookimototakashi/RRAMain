"""
作成者:沖本卓士
作成日:
最終更新日:2022/11/14
稼働設定:解像度 1920*1080 表示スケール125%
####################################################
処理の流れ
####################################################
1:MJSにログインする
↓
2:[デフォルト:\\\\NAS-SV\\B_監査etc\\B2_電子ﾌｧｲﾙ\\RPA_ミロクシステム次年更新\\一括更新申請]
    フォルダ内の[一括更新申請ミロク.xlsm]シートを
    [一括更新申請]フォルダ内の[MJSLog]フォルダへ移動
↓
3:移動したエクセルシートを読取る
↓
4:読取った内容に応じて、MJS各システム年度更新を実行
↓
5:[一括更新申請]フォルダ内の[ミロク更新状況.xlsx]ファイルで
    実行ログを表示する為のテキスト[一括更新申請\\MJSLog\\MJSSysUpLog.txt]を処理毎に出力
↓
6:実行内容に応じてエクセルシートに結果を入力
####################################################
"""

# モジュールインポート
import pyautogui as pg
import time
import pandas as pd
import numpy as np
import os
import subprocess
import traceback
import pyperclip
import logging.config
import datetime
import openpyxl

# 自作モジュールインポート
import WarekiHenkan as WH
import RPA_Function as RPA
import MJSOpen
import Sub_GenkasyoukyakuUpdate as GenkasyoukyakuUpdate
import Sub_HoujinzeiUpdate as HoujinzeiUpdate
import Sub_HouteiUpdate as HouteiUpdate
import Sub_KaikeiUpDate as KaikeiUpDate
import Sub_KessanUpDate as KessanUpDate
import Sub_NencyouUpdate as NencyouUpdate
import Sub_SyotokuzeiUpdate as SyotokuzeiUpdate
import Sub_ZaisanUpdate as ZaisanUpdate

# logger設定------------------------------------------------------------------------------------------------------------
logging.config.fileConfig(r"LogConf\loggingMJSSysUp.conf")
logger = logging.getLogger(__name__)
LURL = r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\一括更新申請\MJSLog\MJSSysUpLog.txt"
# ----------------------------------------------------------------------------------------------------------------------
# #######################################################################################################################
# 一括更新処理の該当年度はThisYearKey.pngなので、年度が変わったらスクリーンショットしなおす事
# #######################################################################################################################

# ------------------------------------------------------------------------------------------------
class Job:
    """
    処理全体(class)
    """

    def __init__(self, **kw):
        log_out("Jobクラス読込開始")
        # 自分のDir(str)
        self.dir = dir
        # 画像のDir(str)
        self.Img_dir = Img_dir
        # 当年(int)
        self.Start_Year = WH.Wareki.from_ad(datetime.datetime.today().year).year
        # RPA用画像フォルダの作成
        self.FolURL = FolURL
        self.TFolURL = TFolURL
        self.imgdir_url = imgdir_url
        self.XLSDir = XLSDir
        self.first_csv = first_csv
        # MJSを起動しログイン後インスタンス化
        self.driver = MJSOpen.MainFlow("self.BatUrl", self.FolURL, self.Img_dir)
        self.TimeOut = False
        if self.driver == "TimeOut":
            log_out("MJSOpen.MainFlowタイムアウト")
            self.TimeOut = True
        else:
            log_out("Jobクラス読込終了")
            self.TimeOut = False

    def KomonUpdate(self, ExRow):
        """
        MJSの顧問先情報更新
        """
        log_out("_顧問先情報更新開始")
        # 関与先コード入力ボックスをクリック------------------------------------
        RPA.ImgClick(self.imgdir_url, r"\Komonsaki_Icon.png", 0.9, 10)
        while (
            pg.locateOnScreen(self.imgdir_url + r"\Komonsaki_Open.png", confidence=0.9)
            is None
        ):
            time.sleep(1)

        p = pg.locateOnScreen(
            self.imgdir_url + r"\Komonsaki_CodeTxt.png", confidence=0.9
        )
        x, y = pg.center(p)
        pg.click(x + 100, y)
        time.sleep(1)
        pg.press("delete")
        pyperclip.copy(str(ExRow["関与先番号"]))
        pg.hotkey("ctrl", "v")
        time.sleep(1)
        pg.press(["return", "return"])

        time.sleep(1)

        p = pg.locateOnScreen(self.imgdir_url + r"\RensaouMeisyou.png", confidence=0.9)
        x, y = pg.center(p)
        pg.click(x + 100, y)
        time.sleep(1)
        pg.press("up")
        pg.press("down")
        pg.press("delete")
        pyperclip.copy(str(ExRow["関与先番号"]))
        pg.hotkey("ctrl", "v")
        time.sleep(1)
        pg.press(["return", "return"])

        pg.keyDown("alt")
        pg.press("u")
        pg.keyUp("alt")

        time.sleep(3)

        pg.keyDown("alt")
        pg.press("x")
        pg.keyUp("alt")

        while (
            pg.locateOnScreen(self.imgdir_url + r"\SyonaiKanri.png", confidence=0.9)
            is None
        ):
            time.sleep(1)

        pg.keyDown("alt")
        pg.press("f4")
        pg.keyUp("alt")
        time.sleep(1)
        log_out("_顧問先情報更新完了")


# ------------------------------------------------------------------------------------------------
class Sheet:
    """
    エクセルブック(class)
    """

    def __init__(self, XLSURL, **kw):
        log_out("_Excelブック読込開始")
        self.mybook_url = XLSURL
        self.book = openpyxl.load_workbook(self.mybook_url, keep_vba=True)
        # 全シート
        self.input_sheet_name = self.book.sheetnames
        # 全シート数
        self.num_sheet = len(self.input_sheet_name)
        log_out("_Excelブック読込終了")

    def Read_sheet(self, sheet_name, first_csv, **kw):
        """
        エクセルシート読込
        """
        log_out("_Excelシート読込開始")

        self.sheet_header = []
        ExSheet = ""
        NameSheet = ""
        ExSheet = self.book[sheet_name]
        ExSheetdata = ExSheet.values
        print(ExSheetdata)
        ExSheetcolumns = next(ExSheetdata)[0:]
        NameSheet = self.book["関与先一覧"]
        NameSheetdata = NameSheet.values
        NameSheetcolumns = next(NameSheetdata)[0:]

        print(ExSheet)
        # 初回読込時の保存--------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H-%M-%S")
        ExSheet = pd.DataFrame(ExSheetdata, columns=ExSheetcolumns)
        self.name_df = pd.DataFrame(NameSheetdata, columns=NameSheetcolumns)

        # 列名整理--------------------------------
        self.sheet_column_count = ExSheet.shape[1]  # 列数
        for Ex in range(self.sheet_column_count):
            ExRow = ExSheet.iloc[0]  # 列名
            ExSecondRow = ExSheet.iloc[1]  # 列名2
            if ExRow[Ex] is None:  # None判定
                # Noneの場合
                Txt = ExRow[Ex - 1]
                if ExSecondRow[Ex] is None:  # None判定
                    # Noneの場合
                    self.sheet_header.append(ExRow[Ex])
                else:
                    # Noneでない場合
                    self.sheet_header.append(Txt + "_" + ExSecondRow[Ex])
            else:
                Txt = ExRow[Ex]
                # Noneでない場合
                if ExSecondRow[Ex] is None:  # None判定
                    # Noneの場合
                    self.sheet_header.append(Txt)
                else:
                    # Noneでない場合
                    self.sheet_header.append(Txt + "_" + ExSecondRow[Ex])
        # データ整理--------------------------------
        self.sheet_header = [e for e in self.sheet_header if e is not None]
        # Df作成
        ExDf = pd.DataFrame(
            ExSheet.values[3:, : len(self.sheet_header)], columns=self.sheet_header
        )
        # Dfnan処理
        ExDf.dropna(how="all", inplace=True)
        print(ExDf)
        self.sheet_df = ExDf
        self.sheet_column_count = self.sheet_df.shape[1]  # 列数
        self.sheet_row_count = self.sheet_df.shape[0]  # 行数
        self.sheet_df.to_csv(
            first_csv + dt_s + ".csv",
            encoding="cp932",
            index=False,
        )
        log_out("_Excelシート読込完了")

    def WriteExcel(self, txt):
        """
        エクセルシート入力
        """
        dt_now = datetime.datetime.now()
        dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
        WriteEx = openpyxl.load_workbook(self.mybook_url, keep_vba=True)
        WriteExSheet = WriteEx[self.sheet_name]
        WriteExSheet.cell(
            row=self.this_row_count + 5, column=self.this_col_count + 2
        ).value = dt_now
        WriteExSheet.cell(
            row=self.this_row_count + 5, column=self.this_col_count + 1
        ).value = txt
        print("シート書き込み完了")
        print(WriteEx)
        WriteEx.save(XLSURL)
        WriteEx.close


# ------------------------------------------------------------------------------------------------
def log_out(txt):
    """
    logger出力
    """
    dt_s = datetime.datetime.now()
    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
    logger.debug(dt_s + txt)


# ------------------------------------------------------------------------------------------------
def logcsv_out(txt):
    """
    loggercsv出力
    """
    dt_s = datetime.datetime.now()
    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
    with open(LURL, "a") as f:
        print([dt_s, txt], file=f)


# ------------------------------------------------------------------------------------------------
def NameSearch(NameDF, Rno):
    """
    顧問先名称の検索
    """
    try:
        NameDFColumn = np.array(NameDF.columns)
        NameDF = np.array(NameDF)
        NC = np.where(NameDFColumn == "コード")
        KC = np.where(NameDFColumn == "顧問先名称")
        NameDFIndex = NameDF[:, NC]
        NR = np.where(NameDFIndex == Rno)
        N_L = NameDF[NR, KC]
        N_L = (
            str(N_L[0])
            .replace("[", "")
            .replace("]", "")
            .replace("'", "")
            .replace('"', "")
            .replace("\u3000", "")
        )
        print(N_L)
        return N_L
    except:
        return "NameErr"


# ------------------------------------------------------------------------------------------------
def ChildFlow_sub(Job, Exc, txt):
    """
    シート列名に応じて処理分岐サブ
    """
    # Log---------------------------------------------------------------------------------------
    msg = "_関与先番号:" + str(Exc.row_kanyo_no) + ":" + str(Exc.row_kanyo_name) + txt + "開始"
    log_out(msg)
    logcsv_out(msg)
    # ------------------------------------------------------------------------------------------
    if "会計大将" == Exc.Title:
        Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
        # 会計大将のアイコンを探す
        ImgList = [r"\K_TaisyouIcon.png", r"\K_TaisyouIcon2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 会計大将のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 会計大将のアイコンをクリック
            ret = KaikeiUpDate.KaikeiUpDate(Job, Exc)
            return ret
    elif "決算内訳書" == Exc.Title:
        Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
        # 決算内訳書のアイコンを探す-------------------------------------------------
        ImgList = [r"\K_Uchiwake.png", r"\K_Uchiwake2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 決算内訳書のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 決算内訳書のアイコンをクリック
            ret = KessanUpDate.KessanUpDate(Job, Exc)
            return ret
    elif "減価償却" == Exc.Title:
        Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
        # 減価償却のアイコンを探す-------------------------------------------------
        ImgList = [r"\G_Syoukyaku.png", r"\G_Syoukyaku2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 減価償却のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 減価償却のアイコンをクリック
            ret = GenkasyoukyakuUpdate.GenkasyoukyakuUpdate(Job, Exc)
            return ret
    elif "法人税申告書" == Exc.Title:
        Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
        # 法人税のアイコンを探す-------------------------------------------------
        ImgList = [r"\Houjinzei.png", r"\Houjinzei2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 法人税のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 法人税のアイコンをクリック
            ret = HoujinzeiUpdate.HoujinzeiUpdate(Job, Exc)
            return ret
    elif "所得税確定申告" == Exc.Title:
        Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
        # 所得税のアイコンを探す-------------------------------------------------
        ImgList = [r"\Syotoku.png", r"\Syotoku2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 所得税のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 所得税のアイコンをクリック
            ret = SyotokuzeiUpdate.SyotokuzeiUpdate(Job, Exc)
            return ret
    elif "財産評価明細書" == Exc.Title:
        Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
        # 財産評価のアイコンを探す-------------------------------------------------
        ImgList = [r"\Zaisanhyouka.png", r"\Zaisanhyouka2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 財産評価のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 財産評価のアイコンをクリック
            ret = ZaisanUpdate.ZaisanUpdate(Job, Exc)
            return ret
    elif "年末調整" == Exc.Title:
        Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
        # 年調のアイコンを探す-------------------------------------------------
        ImgList = [r"\Nencyou.png", r"\Nencyou2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 年調のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 年調のアイコンをクリック
            ret = NencyouUpdate.NencyouUpdate(Job, Exc)
            return ret
    elif "法定調書" == Exc.Title:
        Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
        # 法定調書のアイコンを探す-------------------------------------------------
        ImgList = [r"\Houtei.png", r"\Houtei2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 法定調書のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 法定調書のアイコンをクリック
            ret = HouteiUpdate.HouteiUpdate(Job, Exc)
            return ret
    # Log---------------------------------------------------------------------------------------
    msg = "_関与先番号:" + str(Exc.row_kanyo_no) + ":" + str(Exc.row_kanyo_name) + txt + "終了"
    log_out(msg)
    logcsv_out(msg)
    # ------------------------------------------------------------------------------------------


# ------------------------------------------------------------------------------------------------
def ChildFlow(Job, Exc):
    """
    シート列名に応じて処理分岐
    """
    if "会計大将" == Exc.Title:
        SystemUp = ChildFlow_sub(Job, Exc, "_会計大将更新処理")
        # Excel書き込み--------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel("○")  # シート書き込み
            return True
        elif SystemUp[1] == "当年データ重複エラー":
            Exc.WriteExcel("当年データ重複エラー")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_会計大将当年データ重複エラー"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "TimeOut":
            return False
        else:
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_会計大将更新処理エラー中断"
            )
            log_out(msg)
            logcsv_out(msg)
            return True
    # ------------------------------------------------------------------------------------------
    elif "決算内訳書" == Exc.Title:
        SystemUp = ChildFlow_sub(Job, Exc, "_決算内訳書更新処理")
        # Excel書き込み--------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel("○")
            return True
        elif SystemUp[1] == "Noren":
            Exc.WriteExcel("連動対象無エラー")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_決算内訳書更新処理終了"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "TimeOut":
            return False
    elif "減価償却" == Exc.Title:
        SystemUp = ChildFlow_sub(Job, Exc, "_減価償却更新処理")
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel("○")  # シート書き込み
            return True
        elif SystemUp[1] == "Noren":
            Exc.WriteExcel("決算未確定更新")  # シート書き込み
            return True
        elif SystemUp[1] == "当年データ重複エラー":
            Exc.WriteExcel("減価償却当年データ重複エラー")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_減価償却当年データ重複エラー"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "TimeOut":
            return False
        else:
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_決算未確定減価償却更新処理終了"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
    elif "法人税申告書" == Exc.Title:
        SystemUp = ChildFlow_sub(Job, Exc, "_法人税申告書更新処理")
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel("○")  # シート書き込み
            return True
        elif SystemUp[1] == "要データ再計算":
            Exc.WriteExcel("要データ再計算")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_法人税申告書更新処理終了"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "要申告指定":
            Exc.WriteExcel("要申告指定")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_法人税申告書申告指定無しの為中断"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "TimeOut":
            return False
    elif "所得税確定申告" == Exc.Title:
        SystemUp = ChildFlow_sub(Job, Exc, "_所得税更新処理")
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel("○")  # シート書き込み
            return True
        elif SystemUp[1] == "当年データ重複エラー":
            Exc.WriteExcel("所得税当年データ重複エラー")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_所得税当年データ重複エラー"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "Nocalc":
            Exc.WriteExcel("計算未処理更新")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_所得税更新処理計算未処理で終了"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[0] is False:
            Exc.WriteExcel("関与先無")  # シート書き込み
            return True
        elif SystemUp[1] == "TimeOut":
            return False
    elif "財産評価明細書" == Exc.Title:
        SystemUp = ChildFlow_sub(Job, Exc, "_財産評価明細書更新処理")
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel("○")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_財産評価明細書更新処理終了"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "該当年度有り":
            Exc.WriteExcel("該当年度有り")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_財産評価明細書更新処理終了_該当年度有り"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "TimeOut":
            return False
    elif "年末調整" == Exc.Title:
        SystemUp = ChildFlow_sub(Job, Exc, "_年末調整更新処理")
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel("○")  # シート書き込み
            return True
        elif SystemUp[1] == "該当年度有り":
            Exc.WriteExcel("該当年度有り")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_年末調整更新該当年度有り処理終了"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "TimeOut":
            return False
    elif "法定調書" == Exc.Title:
        SystemUp = ChildFlow_sub(Job, Exc, "_法定調書更新処理")
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel("○")  # シート書き込み
            return True
        elif SystemUp[1] == "該当年度有り":
            Exc.WriteExcel("該当年度有り")  # シート書き込み
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_法定調書更新処理該当年度有り処理終了"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
            return True
        elif SystemUp[1] == "TimeOut":
            return False
    else:
        # Log---------------------------------------------------------------------------------------
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + "_NoSystem終了"
        )
        log_out(msg)
        logcsv_out(msg)
        # ------------------------------------------------------------------------------------------
        print("NoSystem")
        return True


# ------------------------------------------------------------------------------------------------
def OpenSystem(Job, Exc):
    """
    エクセルシート列処理
    """
    try:
        Exc.this_col_count = 0
        for ExrcHeaderItem in Exc.sheet_header:
            if Exc.this_col_count < (len(Exc.sheet_header)):
                if "_繰越対象" in ExrcHeaderItem:
                    SysN = ExrcHeaderItem.split("_")
                    Exc.Title = str(SysN[0])
                    if (
                        not Exc.row_data[Exc.Title + "_繰越対象"] == "-"
                        and str(Exc.row_data[Exc.Title + "_繰越対象"]) == "1"
                    ):
                        if "::" not in Exc.Title:
                            if Exc.row_data[Exc.Title + "_繰越対象"] is None:
                                # Noneの場合
                                if Exc.row_data[Exc.Title + "_繰越処理日"] is None:
                                    # Noneの場合
                                    print(Exc.row_data[Exc.Title + "_繰越処理日"])
                                else:
                                    # Noneでない場合
                                    # Log--------------------------------------------
                                    msg = (
                                        "_関与先番号:"
                                        + str(Exc.row_kanyo_no)
                                        + ":"
                                        + str(Exc.row_kanyo_name)
                                        + "_メイン処理開始"
                                    )
                                    log_out(msg)
                                    logcsv_out(msg)
                                    # -----------------------------------------------
                                    if str(Exc.row_kanyo_no) != "1":
                                        CF = ChildFlow(Job, Exc)
                                        # TimeOut処理
                                        if CF is False:
                                            return False
                            else:
                                # Noneでない場合
                                if Exc.row_data[Exc.Title + "_繰越処理日"] is None:
                                    # Noneの場合
                                    # Log--------------------------------------------
                                    msg = (
                                        "_関与先番号:"
                                        + str(Exc.row_kanyo_no)
                                        + ":"
                                        + str(Exc.row_kanyo_name)
                                        + "_メイン処理開始"
                                    )
                                    log_out(msg)
                                    logcsv_out(msg)
                                    # -----------------------------------------------
                                    if str(Exc.row_kanyo_no) != "1":
                                        CF = ChildFlow(Job, Exc)
                                        # TimeOut処理
                                        if CF is False:
                                            return "TimeOut"
                                else:
                                    # Noneでない場合
                                    print("スタート")
            Exc.this_col_count += 1
        return True
    except:
        print("TEST")
        return False


# ------------------------------------------------------------------------------------------------
def MainStarter(Job, Exc):
    """
    エクセルシート行処理
    """
    try:
        for Exc.this_row_count in range(Exc.sheet_row_count):
            if Exc.this_row_count != 0:
                Exc.row_data = Exc.sheet_df.iloc[Exc.this_row_count]
                if Exc.row_data["関与先番号"] is None:  # nan判定
                    # Noneでない場合
                    print("nan")
                else:
                    # Noneの場合
                    Exc.row_kanyo_no = Exc.row_data["関与先番号"]
                    Exc.row_kanyo_name = NameSearch(Exc.name_df, Exc.row_kanyo_no)
                    OS = OpenSystem(Job, Exc)
                    if OS == "TimeOut":
                        return False, "TimeOut"
                    print("")
        return True, ""
    except:
        return False, ""


# ------------------------------------------------------------------------------------------------
def MainFlow(Job, Exc):
    """
    エクセルブック内各シート処理
    """
    try:
        log_out("xlsxをDataFrameに")
        open(LURL, "w").close()
        for Exc.sheet_name in Exc.input_sheet_name:
            # DataFrameとしてsheetのデータ読込み
            if "更新申請" in Exc.sheet_name:
                Exc.Read_sheet(Exc.sheet_name, Job.first_csv)
                MS = MainStarter(
                    Job,
                    Exc,
                )  # データ送信画面までの関数
                if MS[1] == "TimeOut":
                    return False, "TimeOut"
    except Exception as e:
        log_out(e)
        return False, ""


# ------------------------------------------------------------------------------------------------
def Main():
    """
    概要: メイン処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param Exc : Excel指示シート(obj)
    @return : bool
    """
    global dir, Img_dir
    global FolURL, TFolURL
    global imgdir_url, XLSDir
    global first_csv
    global XLSURL, MoveXLSURL

    dir = RPA.My_Dir("MJS_System_NextCreate")
    Img_dir = dir + r"\\img"
    FolURL = os.getcwd().replace("\\", "/")  # 先
    TFolURL = RPA.My_Dir("MJS_System_NextCreate")  # 先
    imgdir_url = TFolURL + r"\\img"  # 先
    XLSDir = r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\一括更新申請"
    first_csv = XLSDir + r"\MJSLog\MJSSysUpLog.txt"  # 処理状況CSVのURL

    j = Job()
    if j.TimeOut is True:
        while j.TimeOut is False:
            j = Job()
    else:
        # Log--------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_MJSシステム更新開始")
        # -----------------------------------------------
        for curDir, dirs, files in os.walk(j.XLSDir):
            if curDir == j.XLSDir:
                for sb_fileItem in files:
                    print(sb_fileItem)
                    if (
                        "一括更新申請ミロク" in sb_fileItem
                        and not "一括更新申請ミロク(原本).xlsm" == sb_fileItem
                    ):
                        XLSURL = (
                            curDir
                            + r"\\"
                            + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        MoveXLSURL = (
                            curDir
                            + r"\\MJSLog\\"
                            + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        os.rename(XLSURL, MoveXLSURL)
                        MoveXLSURL = (
                            curDir
                            + r"\\"
                            + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        XLSURL = (
                            curDir
                            + r"\\MJSLog\\"
                            + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        Ex_File = Sheet(XLSURL)
                        try:
                            MF = MainFlow(j, Ex_File)
                            # TimeOut処理
                            if MF[1] == "TimeOut":
                                killcmd = "taskkill /F /PID {pid} /T".format(
                                    pid=j.driver.pid
                                )
                                subprocess.run(killcmd, shell=True)
                                log_out("TimeOutによるsubprocess強制終了")
                                del Ex_File  # エクセルブッククラスを解放
                                log_out("Excel解放")
                                os.rename(XLSURL, MoveXLSURL)
                                log_out("Excelファイル移動完了")
                                # TimeOutなら抜ける
                                return False
                        except:
                            traceback.print_exc()
                            del Ex_File  # エクセルブッククラスを解放
                            log_out("Excel解放")
                            os.rename(XLSURL, MoveXLSURL)
                            log_out("Excelファイル移動完了")
        return True


# ------------------------------------------------------------------------------------------------
if __name__ == "__main__":

    while True:  # 無限ループ
        M = Main()
        if M is False:
            M = Main()
        else:
            break
    print("正常終了")
